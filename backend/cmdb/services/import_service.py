# -*- coding: utf-8 -*-
from __future__ import annotations

"""ImportService — 批量导入/导出（CSV / JSON / Excel）

利用 Neo4j UNWIND + MERGE 实现高效的批量写入。
支持增量更新和批量关系创建。
支持 Excel (.xlsx) 格式的导入导出。
"""

import csv
import io
import json
import logging
from uuid import uuid4
from datetime import datetime

from django.core.exceptions import ValidationError

from ..models.model_definition import ModelDefinition
from .neo4j_client import graph_driver
from .validation_service import ValidationService

logger = logging.getLogger(__name__)


def _now_iso() -> str:
    return datetime.now().strftime('%Y-%m-%dT%H:%M:%S')


try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl 未安装，Excel 导入导出功能不可用。安装: pip install openpyxl")


class ImportService:
    """批量导入/导出服务"""

    def import_instances(self, model_code: str,
                         records: list[dict],
                         strategy: str = 'create_or_update') -> dict:
        """批量导入实例

        Args:
            model_code: 目标模型 code
            records: 实例数据列表
            strategy: create_only / create_or_update

        Returns:
            {total: N, created: N, updated: N, errors: [...]}
        """
        model_def = ModelDefinition.objects.get(code=model_code)
        validator = ValidationService(model_def)

        validated_records = []
        errors = []
        now = _now_iso()

        for idx, record in enumerate(records):
            try:
                validated = validator.validate(record)
                validated['__updated_at'] = now
                # 保留系统字段（模型定义中不存在的字段，但 Neo4j MERGE 需要）
                for sys_field in ('instance_id', '__model_code', '__created_at'):
                    if sys_field in record and sys_field not in validated:
                        validated[sys_field] = record[sys_field]
                if not validated.get('instance_id'):
                    validated['instance_id'] = str(uuid4())
                    validated['__model_code'] = model_code
                    validated['__created_at'] = now
                validated_records.append(validated)
            except ValidationError as e:
                errors.append({'index': idx, 'error': str(e)})

        if not validated_records:
            return {'total': 0, 'created': 0, 'updated': 0, 'errors': errors}

        if strategy == 'create_or_update':
            cypher = (
                f"UNWIND $records AS rec "
                f"MERGE (n:{model_code} {{instance_id: rec.instance_id}}) "
                f"SET n += rec "
                f"RETURN count(n) as total"
            )
        else:
            cypher = (
                f"UNWIND $records AS rec "
                f"CREATE (n:{model_code}) "
                f"SET n = rec "
                f"RETURN count(n) as total"
            )

        with graph_driver.session() as session:
            result = session.run(cypher, records=validated_records)
            total = result.single()['total'] if result else 0

        logger.info(f"批量导入 {model_code}: {total} 条记录, {len(errors)} 条错误")
        return {
            'total': total,
            'created': 0,
            'updated': 0,
            'errors': errors,
        }

    def import_relations(self, relations: list[dict]) -> dict:
        if not relations:
            return {'total': 0, 'errors': []}

        now = _now_iso()
        for r in relations:
            r.setdefault('rel_id', str(uuid4()))
            r.setdefault('__created_at', now)

        cypher = (
            "UNWIND $rels AS r "
            "MATCH (src {instance_id: r.src_id}) "
            "MATCH (dst {instance_id: r.dst_id}) "
            "CALL apoc.merge.relationship("
            "  src, r.type, {}, {rel_id: r.rel_id}, dst"
            ") YIELD rel "
            "RETURN count(rel) as total"
        )

        try:
            with graph_driver.session() as session:
                result = session.run(cypher, rels=relations)
                total = result.single()['total'] if result else 0
            return {'total': total, 'errors': []}
        except Exception as e:
            logger.error(f"批量导入关联失败: {e}")
            return {'total': 0, 'errors': [str(e)]}

    def parse_csv(self, csv_content: str) -> list[dict]:
        reader = csv.DictReader(io.StringIO(csv_content))
        return [dict(row) for row in reader]

    def parse_json(self, json_content: str) -> list[dict]:
        data = json.loads(json_content)
        if isinstance(data, list):
            return data
        if isinstance(data, dict) and 'records' in data:
            return data['records']
        raise ValidationError("JSON 格式无效：应为数组或 {records: [...]}")

    def export_to_excel(self, model_code: str, filters: dict = None) -> io.BytesIO:
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl 未安装，请执行: pip install openpyxl")

        model_def = ModelDefinition.objects.get(code=model_code)

        from .node_service import NodeService
        svc = NodeService(model_code)
        result = svc.list(filters or {}, page=1, page_size=999999)
        items = result.get('items', [])

        fields = list(model_def.fields.all().order_by('sort_order', 'name'))
        system_fields = {'instance_id', '__model_code', '__created_at', '__updated_at', 'rel_id'}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = model_def.name or model_code

        headers = [f.label or f.name for f in fields if f.name not in system_fields]
        headers.insert(0, 'instance_id')
        ws.append(headers)

        for item in items:
            row = [item.get('instance_id', '')]
            for f in fields:
                if f.name in system_fields:
                    continue
                val = item.get(f.name)
                if val is None:
                    row.append('')
                elif isinstance(val, (list, dict)):
                    row.append(json.dumps(val, ensure_ascii=False))
                else:
                    row.append(val)
            ws.append(row)

        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                    min_row=2, max_row=min(len(items) + 1, 50)):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 60))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 4

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def import_from_excel(self, model_code: str, file_content: bytes) -> dict:
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl 未安装，请执行: pip install openpyxl")

        model_def = ModelDefinition.objects.get(code=model_code)
        fields = list(model_def.fields.all().order_by('sort_order', 'name'))
        field_map = {f.name: f for f in fields}

        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active

        header_row = [cell.value for cell in ws[1]]
        if not header_row or not header_row[0]:
            raise ValidationError("Excel 文件为空或表头无效")

        col_mapping = {}
        for h in header_row:
            if h == 'instance_id':
                col_mapping[h] = 'instance_id'
            else:
                matched = False
                for f in fields:
                    if f.label == h or f.name == h:
                        col_mapping[h] = f.name
                        matched = True
                        break
                if not matched:
                    col_mapping[h] = h

        records = []
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            record = {}
            for col_idx, header in enumerate(header_row):
                if col_idx < len(row) and row[col_idx] is not None:
                    field_name = col_mapping.get(header, header)
                    record[field_name] = row[col_idx]
            records.append(record)

        return self.import_instances(model_code, records, strategy='create_or_update')
